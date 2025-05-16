import os
import stat
import io
import shutil
import threading
import time
import requests
import openpyxl
from langchain.docstore.document import Document
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import Chroma
from langchain_community.vectorstores.utils import filter_complex_metadata
from langchain.prompts import PromptTemplate
from langchain.schema.runnable import RunnablePassthrough
from langchain_core.output_parsers import StrOutputParser

HF_API_KEY = os.environ.get("HF_API_KEY")
if not HF_API_KEY:
    raise RuntimeError(
        "HuggingFace API key not found. Please set the HF_API_KEY environment variable. "
        "You can get an API key from https://huggingface.co/settings/tokens"
    )

HF_LLM_URL = "https://api-inference.huggingface.co/models/google/flan-t5-small"
HF_EMBED_URL = "https://api-inference.huggingface.co/pipeline/feature-extraction/sentence-transformers/all-MiniLM-L6-v2"

def hf_generate(prompt):
    headers = {"Authorization": f"Bearer {HF_API_KEY}"}
    payload = {"inputs": prompt}
    try:
        response = requests.post(HF_LLM_URL, headers=headers, json=payload)
        response.raise_for_status()
        result = response.json()
        if isinstance(result, list) and "generated_text" in result[0]:
            return result[0]["generated_text"]
        elif isinstance(result, dict) and "generated_text" in result:
            return result["generated_text"]
        else:
            return str(result)
    except requests.exceptions.RequestException as e:
        if response.status_code == 401:
            raise RuntimeError("Invalid HuggingFace API key. Please check your API key at https://huggingface.co/settings/tokens")
        elif response.status_code == 403:
            raise RuntimeError("Access denied. Please check if your HuggingFace API key has the correct permissions.")
        else:
            raise RuntimeError(f"Error calling HuggingFace API: {str(e)}")

def hf_embed(texts):
    headers = {"Authorization": f"Bearer {HF_API_KEY}"}
    payload = {"inputs": texts}
    try:
        response = requests.post(HF_EMBED_URL, headers=headers, json=payload)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        if response.status_code == 401:
            raise RuntimeError("Invalid HuggingFace API key. Please check your API key at https://huggingface.co/settings/tokens")
        elif response.status_code == 403:
            raise RuntimeError("Access denied. Please check if your HuggingFace API key has the correct permissions.")
        else:
            raise RuntimeError(f"Error calling HuggingFace API: {str(e)}")

class ChatData:
    def __init__(self):
        self.persist_directory = "chroma_db"
        os.makedirs(self.persist_directory, exist_ok=True)
        self._ensure_writable(self.persist_directory)

        self.vector_store = None
        self.retriever = None
        self.chain = None
        self.chat_history = []

        self.text_splitter = RecursiveCharacterTextSplitter(chunk_size=1024, chunk_overlap=100)
        self.prompt = PromptTemplate.from_template(
            """
            You are a helpful assistant focused on analyzing Excel data and generating insights. Stick strictly to the provided context. 
            If the question cannot be answered with the given data, respond: "I cannot perform this analysis or the answer is not in the provided data."
            You can return answers as:
            - Plain text
            - Markdown tables (prefixed with "TABLE:")
            - Python code for Streamlit charts (prefixed with "CODE:" and inside triple backticks).
            Question: {question}
            Context: {context}
            Answer:
            """
        )

        self.ping_url = os.environ.get("RENDER_EXTERNAL_URL")
        self.keep_awake = True
        self._start_self_pinger()

    def _ensure_writable(self, path):
        try:
            os.chmod(path, stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO)
            for root, dirs, files in os.walk(path):
                for d in dirs:
                    os.chmod(os.path.join(root, d), stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO)
                for f in files:
                    os.chmod(os.path.join(root, f), stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO)
        except Exception as e:
            raise RuntimeError(f"Failed to set permissions on {path}: {str(e)}")

    def _start_self_pinger(self):
        if not self.ping_url:
            print("Self-ping URL not set. Skipping self-ping setup.")
            return

        def ping():
            while self.keep_awake:
                try:
                    print(f"Pinging {self.ping_url} to prevent Render sleep...")
                    requests.get(self.ping_url, timeout=10)
                except Exception as e:
                    print(f"Self-ping failed: {e}")
                time.sleep(600)  # every 10 minutes

        thread = threading.Thread(target=ping, daemon=True)
        thread.start()

    def _initialize_vector_store(self, documents):
        try:
            # Use HuggingFace embeddings for all documents
            texts = [doc.page_content for doc in documents]
            embeddings = hf_embed(texts)
            for i, doc in enumerate(documents):
                doc.embedding = embeddings[i]
            self.vector_store = Chroma.from_documents(
                documents=documents,
                embedding_function=lambda docs: hf_embed([d.page_content for d in docs]),
                persist_directory=self.persist_directory
            )
            self.vector_store.persist()
        except Exception as e:
            raise RuntimeError(f"Failed to initialize vector store: {str(e)}")

    def _is_excel_file(self, file_obj):
        try:
            current_pos = file_obj.tell()
            try:
                wb = openpyxl.load_workbook(file_obj, read_only=True)
                if not wb.sheetnames:
                    return False
                # Try to read the first sheet to validate it's a proper Excel file
                ws = wb[wb.sheetnames[0]]
                next(ws.iter_rows(values_only=True), None)  # Try to read first row
                return True
            except Exception as e:
                print(f"Excel validation error: {str(e)}")
                return False
            finally:
                file_obj.seek(current_pos)
        except Exception as e:
            print(f"File validation error: {str(e)}")
            return False

    def _validate_excel_data(self, wb):
        """Validate Excel workbook structure and content."""
        if not wb.sheetnames:
            raise ValueError("The Excel file contains no sheets.")
        
        valid_sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            
            # Skip empty sheets
            if not rows:
                continue
                
            # Validate headers
            headers = [str(h).strip() for h in rows[0] if h is not None]
            if not headers:
                continue
                
            # Check for data rows
            has_data = False
            for row in rows[1:]:
                if any(cell is not None for cell in row):
                    has_data = True
                    break
            
            if has_data:
                valid_sheets.append((sheet_name, headers))
        
        if not valid_sheets:
            raise ValueError(
                "No valid data found in the Excel file. Please ensure the file contains:\n"
                "1. At least one sheet\n"
                "2. Headers in the first row\n"
                "3. Data in subsequent rows"
            )
        
        return valid_sheets

    def _process_sheet(self, ws, sheet_name, headers):
        """Process a single sheet and return Document objects."""
        docs = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:  # Skip header row
                continue
                
            # Skip empty rows
            if all(v is None for v in row):
                continue
                
            # Create row data dictionary
            row_data = {}
            for header, value in zip(headers, row):
                if value is not None:
                    # Convert value to string and clean it
                    if isinstance(value, (int, float)):
                        value = str(value).strip()
                    elif isinstance(value, str):
                        value = value.strip()
                    else:
                        value = str(value).strip()
                    row_data[header] = value
            
            if row_data:  # Only add non-empty rows
                content = (
                    f"Sheet: {sheet_name}\n"
                    f"Row {i}\n"
                    f"Data:\n{row_data}"
                )
                docs.append(Document(
                    page_content=content,
                    metadata={
                        "sheet": sheet_name,
                        "row": i,
                        "headers": headers
                    }
                ))
        
        return docs

    def ingest_excel(self, excel_file):
        """Ingest Excel file with improved error handling and data processing."""
        try:
            if not isinstance(excel_file, io.BytesIO):
                raise TypeError("Uploaded file must be a BytesIO object")
            
            if not self._is_excel_file(excel_file):
                raise ValueError(
                    "The uploaded file is not a valid Excel file or is corrupted. "
                    "Please ensure you're uploading a valid .xlsx file."
                )
            
            excel_file.seek(0)
            wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)
            
            # Validate Excel structure and get valid sheets
            valid_sheets = self._validate_excel_data(wb)
            
            # Process each valid sheet
            all_docs = []
            for sheet_name, headers in valid_sheets:
                ws = wb[sheet_name]
                sheet_docs = self._process_sheet(ws, sheet_name, headers)
                all_docs.extend(sheet_docs)
            
            if not all_docs:
                raise ValueError("No valid data could be extracted from the Excel file.")
            
            # Split documents into chunks
            chunks = self.text_splitter.split_documents(all_docs)
            chunks = filter_complex_metadata(chunks)
            
            # Initialize or update vector store
            if self.vector_store is None:
                self._initialize_vector_store(chunks)
            else:
                self.vector_store.add_documents(chunks)
                self.vector_store.persist()
            
            self._create_retriever_and_chain()
            
            return len(all_docs)  # Return number of rows processed
            
        except Exception as e:
            error_msg = str(e)
            if "not a valid Excel file" in error_msg:
                raise ValueError(error_msg)
            elif "No valid data" in error_msg:
                raise ValueError(error_msg)
            else:
                raise RuntimeError(f"Failed to process Excel file: {error_msg}")
        finally:
            try:
                wb.close()
            except:
                pass

    def _create_retriever_and_chain(self):
        if self.vector_store:
            self.retriever = self.vector_store.as_retriever(
                search_type="similarity_score_threshold",
                search_kwargs={"k": 3, "score_threshold": 0.5}
            )
            self.chain = self._chain

    def _chain(self, question):
        # Retrieve context
        docs = self.retriever.get_relevant_documents(question)
        context = "\n".join([doc.page_content for doc in docs])
        prompt = self.prompt.format(question=question, context=context)
        return hf_generate(prompt)

    def ask(self, query: str):
        if not self.chain:
            return "Please, add an Excel file first."
        try:
            response = self.chain(query).strip()
            self.chat_history.append({"question": query, "answer": response})
            if response.startswith("```python") and response.endswith("````"):
                code = response[9:-3].strip()
                return f"CODE:\n{code}"
            elif response.startswith("|") and "\n|" in response:
                return f"TABLE:{response}"
            else:
                return response
        except Exception as e:
            return f"Error processing your question: {str(e)}"

    def clear(self):
        try:
            if os.path.exists(self.persist_directory):
                shutil.rmtree(self.persist_directory)
            self.vector_store = None
            self.retriever = None
            self.chain = None
            self.chat_history = []
        except Exception as e:
            raise RuntimeError(f"Failed to clear data: {str(e)}")

    def export_chat_history(self):
        if not self.chat_history:
            return "No chat history available."
        try:
            lines = []
            for entry in self.chat_history:
                lines.append(f"Q: {entry['question']}\nA: {entry['answer']}\n")
            return "\n".join(lines)
        except Exception as e:
            raise RuntimeError(f"Failed to export chat history: {str(e)}")
